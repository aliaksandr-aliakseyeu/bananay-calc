"""
Скрипт импорта точек доставки из Excel файла в базу данных.

Использование:
    poetry run python scripts/import_delivery_points.py
    poetry run python scripts/import_delivery_points.py --debug
"""
import argparse
import asyncio
import io
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select, text
from tqdm import tqdm

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.db.base import engine
from app.db.models import (District, Settlement, Tag,
                           delivery_point_tags)
from app.db.models.enums import SettlementType
from app.utils.slugify import slugify

BATCH_SIZE = 100
DEBUG_MODE = False


class ValidationError(Exception):
    """Ошибка валидации данных строки (не SQL ошибка)."""
    pass


SETTLEMENT_TYPE_MAP = {
    'пгт': SettlementType.PGT,
    'г.': SettlementType.GOROD,
    'г': SettlementType.GOROD,
    'город': SettlementType.GOROD,
    'село': SettlementType.SELO,
    'с.': SettlementType.SELO,
    'деревня': SettlementType.DEREVNYA,
    'дер.': SettlementType.DEREVNYA,
    'д.': SettlementType.DEREVNYA,
    'поселок': SettlementType.POSELOK,
    'пос.': SettlementType.POSELOK,
    'п.': SettlementType.POSELOK,
    'станица': SettlementType.STANITSA,
    'ст.': SettlementType.STANITSA,
    'хутор': SettlementType.KHUTOR,
    'х.': SettlementType.KHUTOR,
    'аул': SettlementType.AUL,
}


class ImportStats:
    """Класс для сбора статистики импорта."""

    def __init__(self):
        self.total_rows = 0
        self.imported_count = 0
        self.error_count = 0
        self.batch_count = 0
        self.errors_by_type = defaultdict(int)
        self.created_settlements = 0
        self.created_districts = 0
        self.created_tags = 0

    def add_error(self, error_type: str):
        """Добавить ошибку."""
        self.error_count += 1
        self.errors_by_type[error_type] += 1

    def print_summary(self):
        """Вывести итоговую статистику."""
        print(f'\n{"="*70}')
        print('📊 ИТОГОВАЯ СТАТИСТИКА ИМПОРТА')
        print(f'{"="*70}')
        print(f'Всего строк:              {self.total_rows}')
        print(f'Успешно импортировано:    {self.imported_count} ({self.imported_count/self.total_rows*100:.1f}%)')
        print(f'Ошибок:                   {self.error_count} ({self.error_count/self.total_rows*100:.1f}%)')
        print(f'Батчей обработано:        {self.batch_count}')
        print('\nСоздано новых записей:')
        print(f'  Населённых пунктов:     {self.created_settlements}')
        print(f'  Районов:                {self.created_districts}')
        print(f'  Тегов:                  {self.created_tags}')

        if self.errors_by_type:
            print('\nОшибки по типам:')
            for error_type, count in sorted(self.errors_by_type.items(), key=lambda x: -x[1]):
                print(f'  {error_type}: {count}')

        print(f'{"="*70}\n')


def parse_settlement(text: str) -> tuple[str, SettlementType]:
    """
    Парсит название населённого пункта и определяет его тип.

    Args:
        text: Строка вида "пгт Джубга" или "Сочи"

    Returns:
        (название, тип)
    """
    if not text:
        raise ValueError("Название населённого пункта пустое")

    text = text.strip()

    for abbr, settlement_type in SETTLEMENT_TYPE_MAP.items():
        pattern = rf'^{re.escape(abbr)}[\s.]+'
        if re.match(pattern, text, re.IGNORECASE):
            name = re.sub(pattern, '', text, flags=re.IGNORECASE).strip()
            return (name, settlement_type)

    return (text, SettlementType.GOROD)


async def get_or_create_settlement(
    conn,
    region_id: int,
    name: str,
    settlement_type: SettlementType,
    postal_code: int | None = None,
    stats: ImportStats = None
) -> int:
    """
    Получает или создаёт населённый пункт.

    Returns:
        ID населённого пункта
    """
    result = await conn.execute(
        select(Settlement.id).where(
            Settlement.region_id == region_id,
            Settlement.name == name,
            Settlement.type == settlement_type
        )
    )
    settlement_id = result.scalar_one_or_none()

    if settlement_id:
        return settlement_id

    result = await conn.execute(
        Settlement.__table__.insert().values(
            region_id=region_id,
            name=name,
            type=settlement_type,
            postal_code=postal_code
        ).returning(Settlement.id)
    )
    settlement_id = result.scalar_one()

    if stats:
        stats.created_settlements += 1

    return settlement_id


async def get_or_create_district(
    conn,
    settlement_id: int,
    name: str,
    stats: ImportStats = None
) -> int:
    """
    Получает или создаёт район.

    Returns:
        ID района
    """
    result = await conn.execute(
        select(District.id).where(
            District.settlement_id == settlement_id,
            District.name == name
        )
    )
    district_id = result.scalar_one_or_none()

    if district_id:
        return district_id

    result = await conn.execute(
        District.__table__.insert().values(
            settlement_id=settlement_id,
            name=name
        ).returning(District.id)
    )
    district_id = result.scalar_one()

    if stats:
        stats.created_districts += 1

    return district_id


async def find_category(conn, name: str) -> int | None:
    """
    Ищет категорию по названию (без учёта регистра).

    Returns:
        ID категории или None
    """
    result = await conn.execute(
        text("SELECT id FROM categories WHERE LOWER(name) = LOWER(:name)"),
        {"name": name}
    )
    return result.scalar_one_or_none()


async def find_subcategory(conn, category_id: int, name: str) -> int | None:
    """
    Ищет подкатегорию по категории и названию (без учёта регистра).

    Returns:
        ID подкатегории или None
    """
    result = await conn.execute(
        text("""
            SELECT id FROM subcategories
            WHERE category_id = :category_id
              AND LOWER(name) = LOWER(:name)
        """),
        {"category_id": category_id, "name": name}
    )
    return result.scalar_one_or_none()


async def get_or_create_tag(conn, name: str, stats: ImportStats = None) -> int:
    """
    Получает или создаёт тег.

    Returns:
        ID тега
    """
    result = await conn.execute(
        text("SELECT id FROM tags WHERE LOWER(name) = LOWER(:name)"),
        {"name": name}
    )
    tag_id = result.scalar_one_or_none()

    if tag_id:
        return tag_id

    slug = slugify(name)
    result = await conn.execute(
        Tag.__table__.insert().values(
            name=name,
            slug=slug
        ).returning(Tag.id)
    )
    tag_id = result.scalar_one()

    if stats:
        stats.created_tags += 1

    return tag_id


def safe_str(value) -> str | None:
    """
    Безопасное преобразование значения из Excel в строку.

    Args:
        value: Значение из Excel (может быть str, int, float, None)

    Returns:
        Строка или None
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return str(value).strip()


def process_tags(rubrika_str: str | None) -> list[str]:
    """
    Парсит строку с тегами, разделёнными точкой с запятой.

    Args:
        rubrika_str: "Быстрое питание; Кафе; Мороженое"

    Returns:
        Список названий тегов
    """
    if not rubrika_str:
        return []

    tags = []
    for tag_name in rubrika_str.split(';'):
        tag_name = tag_name.strip()
        if tag_name:
            tags.append(tag_name)

    return tags


async def process_single_row(
    conn,
    row_num: int,
    row_data: dict,
    region_id: int,
    stats: ImportStats
) -> bool:
    """
    Обрабатывает одну строку Excel.

    Returns:
        True если успешно, False если ошибка
    """
    try:
        if not row_data.get('Название'):
            raise ValidationError('Нет названия')

        if not row_data.get('X') or not row_data.get('Y'):
            raise ValidationError('Нет координат')

        settlement_text = row_data.get('Населенный Пункт')
        if not settlement_text:
            raise ValidationError('Нет населённого пункта')

        settlement_name, settlement_type = parse_settlement(settlement_text)

        postal_code = row_data.get('Индекс')
        settlement_id = await get_or_create_settlement(
            conn, region_id, settlement_name, settlement_type, postal_code, stats
        )

        district_id = None
        district_text = row_data.get('Район')
        if district_text and district_text.strip():
            district_id = await get_or_create_district(
                conn, settlement_id, district_text.strip(), stats
            )

        category_name = row_data.get('Раздел')
        if not category_name:
            raise ValidationError('Нет раздела')

        category_id = await find_category(conn, category_name)
        if not category_id:
            raise ValidationError(f'Категория "{category_name}" не найдена')

        subcategory_name = row_data.get('Подраздел')
        if not subcategory_name:
            raise ValidationError('Нет подраздела')

        subcategory_id = await find_subcategory(conn, category_id, subcategory_name)
        if not subcategory_id:
            raise ValidationError(f'Подкатегория "{subcategory_name}" не найдена')

        tag_names = process_tags(row_data.get('Рубрика'))
        tag_ids = []
        for tag_name in tag_names:
            tag_id = await get_or_create_tag(conn, tag_name, stats)
            tag_ids.append(tag_id)

        longitude = float(row_data['Y'])
        latitude = float(row_data['X'])

        result = await conn.execute(
            text("""
                INSERT INTO delivery_points (
                    name, type, title,
                    settlement_id, district_id,
                    address, address_comment, landmark,
                    location,
                    category_id, subcategory_id,
                    phone, mobile, email, schedule,
                    is_active
                ) VALUES (
                    :name, :type, :title,
                    :settlement_id, :district_id,
                    :address, :address_comment, :landmark,
                    ST_GeomFromText(:location, 4326),
                    :category_id, :subcategory_id,
                    :phone, :mobile, :email, :schedule,
                    :is_active
                ) RETURNING id
            """),
            {
                "name": safe_str(row_data.get('Название')),
                "type": safe_str(row_data.get('Тип')),
                "title": safe_str(row_data.get('Заголовок')),
                "settlement_id": settlement_id,
                "district_id": district_id,
                "address": safe_str(row_data.get('Адрес')),
                "address_comment": safe_str(row_data.get('Комментарий к адресу')),
                "landmark": safe_str(row_data.get('Метро (Остановка)')),
                "location": f'POINT({longitude} {latitude})',
                "category_id": category_id,
                "subcategory_id": subcategory_id,
                "phone": safe_str(row_data.get('Телефоны')),
                "mobile": safe_str(row_data.get('Сотовый')),
                "email": safe_str(row_data.get('Email')),
                "schedule": safe_str(row_data.get('Время Работы')),
                "is_active": True
            }
        )
        delivery_point_id = result.scalar_one()

        for tag_id in tag_ids:
            await conn.execute(
                delivery_point_tags.insert().values(
                    delivery_point_id=delivery_point_id,
                    tag_id=tag_id
                )
            )

        return True

    except ValidationError as e:
        stats.add_error(str(e))
        return False

    except Exception as e:
        if DEBUG_MODE:
            print(f'\n{"="*70}')
            print(f'🐛 DEBUG: SQL ОШИБКА В СТРОКЕ EXCEL {row_num + 1}')
            print(f'{"="*70}')
            print('\nДанные строки:')
            for key, value in row_data.items():
                if value:
                    print(f'  {key}: {value}')
            print('\nТекст ошибки:')
            print(f'  {e}')
            print(f'\n{"="*70}\n')
            import traceback
            traceback.print_exc()
            raise
        else:
            raise


async def import_delivery_points(excel_path: str, region_id: int = 1):
    """
    Импортирует точки доставки из Excel файла с батчами.

    Args:
        excel_path: Путь к Excel файлу
        region_id: ID региона (по умолчанию 1 - Краснодарский край)
    """
    stats = ImportStats()

    print(f'📂 Открываю файл: {Path(excel_path).name}')
    wb = load_workbook(excel_path)
    ws = wb.active

    columns = [cell.value for cell in ws[1]]
    stats.total_rows = ws.max_row - 1

    print(f'📊 Найдено строк: {stats.total_rows}')
    print(f'⚙️  Размер батча: {BATCH_SIZE} строк')
    if DEBUG_MODE:
        print('🐛 РЕЖИМ ОТЛАДКИ: Выполнение остановится на первой SQL ошибке')
    print('🚀 Начинаю импорт...\n')

    async with engine.connect() as conn:
        result = await conn.execute(
            text("SELECT name FROM regions WHERE id = :region_id"),
            {"region_id": region_id}
        )
        region_row = result.fetchone()

        if not region_row:
            raise ValueError(f'Регион с ID {region_id} не найден')

        print(f'✓ Регион: {region_row[0]} (ID: {region_id})\n')

    total_batches = (stats.total_rows + BATCH_SIZE - 1) // BATCH_SIZE

    with tqdm(total=stats.total_rows, desc="Импорт", unit="строк") as pbar:
        for batch_start in range(2, ws.max_row + 1, BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, ws.max_row + 1)
            batch_num = stats.batch_count + 1
            batch_size = batch_end - batch_start

            batch_success = 0
            current_excel_row = batch_start

            try:
                async with engine.begin() as conn:
                    for row_idx in range(batch_start, batch_end):
                        current_excel_row = row_idx
                        row_num = row_idx - 1

                        row_data = {}
                        for col_idx, col_name in enumerate(columns, 1):
                            cell_value = ws.cell(row_idx, col_idx).value
                            row_data[col_name] = cell_value

                        success = await process_single_row(
                            conn, row_num, row_data, region_id, stats
                        )

                        if success:
                            stats.imported_count += 1
                            batch_success += 1

                        pbar.update(1)

                stats.batch_count += 1

                tqdm.write(f'  ✓ Батч {batch_num}/{total_batches}: '
                           f'успешно {batch_success}/{batch_size}')

            except Exception as e:
                if DEBUG_MODE:
                    raise
                else:
                    remaining_rows = batch_size - batch_success
                    tqdm.write(
                        f'  ✗ Батч {batch_num}/{total_batches}: SQL ОШИБКА в строке Excel {current_excel_row}'
                    )
                    tqdm.write(f'     Детали: {str(e)[:150]}')
                    tqdm.write(f'     Пропущено строк из батча: {remaining_rows}')
                    stats.error_count += remaining_rows
                    stats.add_error(f'SQL ошибка: {str(e)[:50]}')
                    pbar.update(remaining_rows)
                    continue

    await engine.dispose()

    stats.print_summary()


async def main():
    """Главная функция."""
    global DEBUG_MODE

    parser = argparse.ArgumentParser(
        description='Импорт точек доставки из Excel в базу данных'
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help='Режим отладки: остановка на первой SQL ошибке с детальной информацией'
    )
    parser.add_argument(
        '--file',
        type=str,
        help='Путь к Excel файлу (по умолчанию: sochi_address.xlsx)'
    )
    parser.add_argument(
        '--region-id',
        type=int,
        default=1,
        help='ID региона (по умолчанию: 1 - Краснодарский край)'
    )
    args = parser.parse_args()

    DEBUG_MODE = args.debug

    if args.file:
        excel_path = Path(args.file)
    else:
        excel_path = Path(__file__).parent.parent / 'sochi_address.xlsx'

    if not excel_path.exists():
        print(f'❌ Файл не найден: {excel_path}')
        sys.exit(1)

    try:
        await import_delivery_points(str(excel_path), region_id=args.region_id)
    except Exception as e:
        if not DEBUG_MODE:
            print(f'\n❌ Критическая ошибка: {e}')
            import traceback
            traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    asyncio.run(main())
